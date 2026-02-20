"""
auto_actuary.reports.renderers.excel
======================================
Excel workbook rendering using openpyxl.

Provides a styled workbook writer used by all actuarial and executive reports.
Style themes follow professional actuarial exhibit conventions:
  - Navy header rows
  - Light blue alternating data rows
  - Bold totals row
  - Right-aligned numbers with commas / % formatting
  - Freeze top row for navigation
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import numpy as np
import pandas as pd

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Color palette
# ---------------------------------------------------------------------------

NAVY      = "002060"
TEAL      = "0070C0"
LIGHT_BLUE = "DEEAF1"
WHITE     = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
GOLD      = "FFC000"
RED       = "C00000"
GREEN     = "375623"
GREEN_LIGHT = "E2EFDA"

# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _font(bold=False, color=WHITE, size=10):
    return Font(bold=bold, color=color, size=size)

def _fill(hex_color: str):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def _border_thin():
    side = Side(style="thin", color="BFBFBF")
    return Border(left=side, right=side, top=side, bottom=side)

def _number_fmt(fmt: str):
    return fmt


# ---------------------------------------------------------------------------
# Main writer class
# ---------------------------------------------------------------------------

class ExcelWriter:
    """
    Write one or more DataFrames to a professionally styled Excel workbook.

    Usage
    -----
    >>> writer = ExcelWriter(primary_color="002060", accent_color="0070C0")
    >>> writer.add_sheet("Triangle", triangle_df, number_format="$#,##0")
    >>> writer.add_sheet("Summary", summary_df)
    >>> writer.save("output/triangle.xlsx")
    """

    def __init__(
        self,
        title: str = "auto_actuary Report",
        company: str = "P&C Carrier",
        primary_color: str = NAVY,
        accent_color: str = TEAL,
    ) -> None:
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for Excel output.  pip install openpyxl")
        self.title = title
        self.company = company
        self.primary_color = primary_color
        self.accent_color = accent_color
        self.wb = Workbook()
        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

    def add_sheet(
        self,
        sheet_name: str,
        df: pd.DataFrame,
        title: Optional[str] = None,
        subtitle: Optional[str] = None,
        number_format: str = "#,##0",
        pct_cols: Optional[List[str]] = None,
        currency_cols: Optional[List[str]] = None,
        highlight_total_row: bool = True,
        freeze_top: int = 1,
        col_widths: Optional[Dict[str, int]] = None,
    ) -> None:
        """
        Add a DataFrame as a styled worksheet.

        Parameters
        ----------
        sheet_name : str
        df : pd.DataFrame
        title : str, optional — large title row above data
        subtitle : str, optional — smaller subtitle row
        number_format : str — default number format for numeric cells
        pct_cols : list[str] — columns to format as percentage
        currency_cols : list[str] — columns to format as $#,##0
        highlight_total_row : bool — bold + fill last row if labelled TOTAL
        freeze_top : int — freeze this many rows from top (1 = header only)
        """
        pct_cols = set(pct_cols or [])
        currency_cols = set(currency_cols or [])

        ws = self.wb.create_sheet(title=sheet_name[:31])

        row_offset = 1

        # Title rows
        if title:
            ws.cell(row=row_offset, column=1).value = title
            ws.cell(row=row_offset, column=1).font = Font(
                bold=True, size=14, color=WHITE
            )
            ws.cell(row=row_offset, column=1).fill = _fill(self.primary_color)
            ws.merge_cells(
                start_row=row_offset, start_column=1,
                end_row=row_offset, end_column=max(len(df.columns) + df.index.nlevels, 5)
            )
            row_offset += 1

        if subtitle:
            ws.cell(row=row_offset, column=1).value = subtitle
            ws.cell(row=row_offset, column=1).font = Font(bold=False, size=10, color=WHITE)
            ws.cell(row=row_offset, column=1).fill = _fill(self.accent_color)
            ws.merge_cells(
                start_row=row_offset, start_column=1,
                end_row=row_offset, end_column=max(len(df.columns) + df.index.nlevels, 5)
            )
            row_offset += 1

        header_row = row_offset

        # Reset index to make it a regular column
        df_out = df.reset_index()
        columns = list(df_out.columns)
        n_cols = len(columns)

        # Header row
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = str(col_name).replace("_", " ").title()
            cell.font = Font(bold=True, color=WHITE, size=10)
            cell.fill = _fill(self.primary_color)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        row_offset = header_row + 1

        # Data rows
        for row_num, (_, row_data) in enumerate(df_out.iterrows()):
            is_total = str(row_data.iloc[0]).upper() in ("TOTAL", "SUBTOTAL", "GRAND TOTAL")
            alt_fill = _fill(LIGHT_BLUE) if row_num % 2 == 0 else _fill(WHITE)
            total_fill = _fill(GOLD)

            for col_idx, (col_name, val) in enumerate(zip(columns, row_data), start=1):
                cell = ws.cell(row=row_offset + row_num, column=col_idx)

                if val is None or (isinstance(val, float) and np.isnan(val)):
                    cell.value = "—"
                elif isinstance(val, (int, float)):
                    cell.value = float(val)
                    # Apply number format
                    if col_name in pct_cols:
                        cell.number_format = "0.0%"
                    elif col_name in currency_cols:
                        cell.number_format = '$#,##0'
                    elif isinstance(val, float) and abs(val) < 10:
                        cell.number_format = "0.000"
                    else:
                        cell.number_format = number_format
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.value = str(val)
                    cell.alignment = Alignment(horizontal="left")

                if is_total and highlight_total_row:
                    cell.fill = total_fill
                    cell.font = Font(bold=True, size=10)
                else:
                    cell.fill = alt_fill

        # Auto column widths
        for col_idx, col_name in enumerate(columns, start=1):
            max_len = max(len(str(col_name)), 8)
            if col_widths and col_name in col_widths:
                max_len = col_widths[col_name]
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)

        # Freeze
        freeze_cell = f"A{header_row + freeze_top}"
        ws.freeze_panes = freeze_cell

    def add_cover(self, produced_by: str = "auto_actuary", as_of_date: Optional[str] = None) -> None:
        """Add a professional cover sheet."""
        ws = self.wb.create_sheet(title="Cover", index=0)
        ws.sheet_view.showGridLines = False

        # Background
        for row in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=15):
            for cell in row:
                cell.fill = _fill(NAVY)

        # Company name
        ws.merge_cells("B5:N5")
        ws["B5"].value = self.company
        ws["B5"].font = Font(bold=True, size=24, color=WHITE)
        ws["B5"].alignment = Alignment(horizontal="left", vertical="center")

        # Report title
        ws.merge_cells("B7:N7")
        ws["B7"].value = self.title
        ws["B7"].font = Font(bold=True, size=18, color=GOLD)
        ws["B7"].alignment = Alignment(horizontal="left", vertical="center")

        # Subtitle
        ws.merge_cells("B9:N9")
        ws["B9"].value = f"Produced by {produced_by}"
        ws["B9"].font = Font(size=11, color=LIGHT_BLUE)

        if as_of_date:
            ws.merge_cells("B10:N10")
            ws["B10"].value = f"As of: {as_of_date}"
            ws["B10"].font = Font(size=11, color=LIGHT_BLUE)

        ws.column_dimensions["A"].width = 3
        for col in "BCDEFGHIJKLMN":
            ws.column_dimensions[col].width = 10

    def save(self, path: Union[str, Path]) -> Path:
        """Save the workbook and return the file path."""
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(path)
        logger.info("Excel workbook saved: %s", path)
        return path
